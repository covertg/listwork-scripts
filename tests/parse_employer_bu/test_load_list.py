"""Tests for `parse_employer_bu.load_list`."""

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from parse_employer_bu import load_list


def _write_xlsx(
    path: Path, df: pd.DataFrame, zip_col: str | None = "ZIP"
) -> Path:
    """Write `df` to `path` using openpyxl, forcing the zip column to Text format
    so leading zeros survive round-tripping through Excel."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="BU")
        if zip_col and zip_col in df.columns:
            ws = writer.sheets["BU"]
            col_idx = df.columns.get_loc(zip_col) + 1
            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
    return path


def test_load_list_standard(tmp_path):
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame(
            {
                "NAME": ["Smith, John", "Doe, Jane"],
                "PROGRAM": ["Computer Science", "Biology"],
            }
        ),
    )
    df = load_list(p)
    assert df.shape == (2, 2)
    assert list(df.columns) == ["NAME", "PROGRAM"]


def test_load_list_nonexistent_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_list(tmp_path / "does_not_exist.xlsx")


def test_load_list_zip_as_string(tmp_path):
    # Regression guard: zip codes with leading zeros must come back as strings.
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame({"NAME": ["Smith, John"], "ZIP": ["03755"]}),
    )
    df = load_list(p)
    assert df["ZIP"].iloc[0] == "03755"
    assert isinstance(df["ZIP"].iloc[0], str)


def test_load_list_strips_whitespace(tmp_path):
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame({"NAME": ["  Smith, John  "], "PROGRAM": [" Biology "]}),
    )
    df = load_list(p)
    assert df["NAME"].iloc[0] == "Smith, John"
    assert df["PROGRAM"].iloc[0] == "Biology"


def test_load_list_empty_string_raises(tmp_path):
    # pd.read_excel with dtype=str turns truly-empty cells into NaN, not "". To
    # trigger the load_list empty-string check we need a cell that contains
    # whitespace and therefore strips to "".
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame({"NAME": ["Smith, John"], "PROGRAM": [" "]}),
    )
    with pytest.raises(RuntimeError):
        load_list(p)


def test_load_list_tolerates_nan(tmp_path):
    # NaN in non-required cells (e.g. ADDRESS_LINE2) should load without error.
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame(
            {
                "NAME": ["Smith, John", "Doe, Jane"],
                "ADDRESS_LINE2": ["Apt 4", np.nan],
            }
        ),
    )
    df = load_list(p)
    assert len(df) == 2
    assert df["ADDRESS_LINE2"].isna().any()


def test_load_list_sample_fixture(sample_xlsx_path):
    """Smoke test against the committed sample fixture."""
    df = load_list(sample_xlsx_path)
    expected_cols = {
        "EMPLID",
        "NAME",
        "PROGRAM",
        "ADDRESS_LINE1",
        "ADDRESS_LINE2",
        "TOWN/CITY",
        "ST",
        "ZIP",
    }
    assert expected_cols.issubset(df.columns)
    assert len(df) == 10  # matches generate_sample.ROWS
    # Zip leading-zero guard, end-to-end via the fixture.
    assert df["ZIP"].iloc[0] == "03755"


def test_load_list_does_not_mutate_file(tmp_path):
    p = _write_xlsx(
        tmp_path / "BU 2024.09.15.xlsx",
        pd.DataFrame({"NAME": ["Smith, John"], "PROGRAM": ["Biology"]}),
    )
    mtime_before = p.stat().st_mtime_ns
    _ = load_list(p)
    # Extra confidence: the xlsx file should not be modified by a read.
    wb = load_workbook(p, read_only=True)
    wb.close()
    assert p.stat().st_mtime_ns == mtime_before
